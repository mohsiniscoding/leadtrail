"""
Full Export Excel Module.

This module provides functionality to export comprehensive campaign data
to Excel format with multiple sheets for complete analysis and reporting.
"""
from datetime import datetime
from io import BytesIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from django.core.exceptions import ObjectDoesNotExist
from leadtrail.portal.models import Campaign, CompanyNumber


def generate_full_export_excel(campaign: Campaign) -> HttpResponse:
    """
    Generate a comprehensive Excel export of all campaign data.
    
    Args:
        campaign: The campaign to export data for
        
    Returns:
        HttpResponse with Excel data as attachment
    """
    # Query all data once with all necessary relationships
    companies = CompanyNumber.objects.filter(
        campaign=campaign
    ).select_related('house_data', 'vat_lookup', 'website_hunting_result', 'website_contact_lookup', 'linkedin_lookup').order_by('company_number')
    
    # Create workbook
    workbook = Workbook()
    
    # Remove default sheet
    workbook.remove(workbook.active)
    
    # Create Companies House sheet
    create_companies_house_sheet(workbook, companies)
    
    # Create VAT Lookup sheet
    create_vat_lookup_sheet(workbook, companies)
    
    # Create Website Hunting sheet
    create_website_hunting_sheet(workbook, companies)
    
    # Create Website Contact Extraction sheet
    create_website_contact_extraction_sheet(workbook, companies)
    
    # Create LinkedIn Profile Discovery sheet
    create_linkedin_profile_discovery_sheet(workbook, companies)
    
    # Save workbook to BytesIO
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create HTTP response with Excel content
    response = HttpResponse(
        excel_buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Format filename with datetime
    filename = f"{campaign.name}-full-export-{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response


def create_companies_house_sheet(workbook: Workbook, companies) -> None:
    """
    Create Companies House Lookup sheet with all company data.
    
    Args:
        workbook: The Excel workbook to add sheet to
        companies: QuerySet of CompanyNumber objects with related data
    """
    # Create worksheet
    sheet = workbook.create_sheet("Companies House Lookup")
    
    # Define headers
    headers = [
        'Company Number',
        'Company Name',
        'Company Status', 
        'Company Type',
        'Incorporation Date',
        'Jurisdiction',
        'Address Line 1',
        'Address Line 2',
        'Locality',
        'Region',
        'Postal Code',
        'Country',
        'Registered Office Address',
        'SIC Codes',
        'Can File',
        'Has Been Liquidated',
        'Has Charges',
        'Has Insolvency History',
        'Last Accounts Date',
        'Next Accounts Due',
        'Accounts Overdue',
        'Confirmation Statement Date',
        'Confirmation Statement Next Due',
        'Confirmation Statement Overdue',
        'Officers Total Count',
        'Officers Active Count',
        'Key Officers',
        'Error Message',
        'Lookup Status',
        'Created At'
    ]
    
    # Add headers to sheet
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row_num, company in enumerate(companies, 2):
        try:
            house_data = company.house_data
        except ObjectDoesNotExist:
            house_data = None
        
        # Prepare row data with fallback to empty strings
        row_data = [
            company.company_number,
            getattr(house_data, 'company_name', '') if house_data else '',
            getattr(house_data, 'company_status', '') if house_data else '',
            getattr(house_data, 'company_type', '') if house_data else '',
            getattr(house_data, 'incorporation_date', '') if house_data else '',
            getattr(house_data, 'jurisdiction', '') if house_data else '',
            getattr(house_data, 'address_line_1', '') if house_data else '',
            getattr(house_data, 'address_line_2', '') if house_data else '',
            getattr(house_data, 'locality', '') if house_data else '',
            getattr(house_data, 'region', '') if house_data else '',
            getattr(house_data, 'postal_code', '') if house_data else '',
            getattr(house_data, 'country', '') if house_data else '',
            getattr(house_data, 'registered_office_address', '') if house_data else '',
            getattr(house_data, 'sic_codes', '') if house_data else '',
            getattr(house_data, 'can_file', '') if house_data else '',
            getattr(house_data, 'has_been_liquidated', '') if house_data else '',
            getattr(house_data, 'has_charges', '') if house_data else '',
            getattr(house_data, 'has_insolvency_history', '') if house_data else '',
            getattr(house_data, 'last_accounts_date', '') if house_data else '',
            getattr(house_data, 'next_accounts_due', '') if house_data else '',
            getattr(house_data, 'accounts_overdue', '') if house_data else '',
            getattr(house_data, 'confirmation_statement_date', '') if house_data else '',
            getattr(house_data, 'confirmation_statement_next_due', '') if house_data else '',
            getattr(house_data, 'confirmation_statement_overdue', '') if house_data else '',
            getattr(house_data, 'officers_total_count', '') if house_data else '',
            getattr(house_data, 'officers_active_count', '') if house_data else '',
            getattr(house_data, 'key_officers', '') if house_data else '',
            getattr(house_data, 'error_message', '') if house_data else '',
            getattr(house_data, 'status', 'NOT_PROCESSED') if house_data else 'NOT_PROCESSED',
            house_data.created_at.strftime('%Y-%m-%d %H:%M:%S') if house_data and house_data.created_at else ''
        ]
        
        # Write row data to sheet
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num).value = value
    
    # Auto-adjust column widths
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)


def create_vat_lookup_sheet(workbook: Workbook, companies) -> None:
    """
    Create VAT Lookup sheet with all VAT lookup data.
    
    Args:
        workbook: The Excel workbook to add sheet to
        companies: QuerySet of CompanyNumber objects with related data
    """
    # Create worksheet
    sheet = workbook.create_sheet("VAT Lookup")
    
    # Define headers
    headers = [
        'Company Number',
        'Company Name (from Companies House)',
        'VAT Number',
        'VAT Company Name',
        'Search Terms',
        'Status',
        'Processing Notes',
        'Proxy Used',
        'Created At'
    ]
    
    # Add headers to sheet
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row_num, company in enumerate(companies, 2):
        try:
            vat_lookup = company.vat_lookup
        except ObjectDoesNotExist:
            vat_lookup = None
        try:
            house_data = company.house_data
        except ObjectDoesNotExist:
            house_data = None
        
        # Prepare row data with fallback to empty strings
        row_data = [
            company.company_number,
            getattr(house_data, 'company_name', '') if house_data else '',
            getattr(vat_lookup, 'vat_number', '') if vat_lookup else '',
            getattr(vat_lookup, 'company_name', '') if vat_lookup else '',
            getattr(vat_lookup, 'search_terms', '') if vat_lookup else '',
            getattr(vat_lookup, 'status', 'NOT_PROCESSED') if vat_lookup else 'NOT_PROCESSED',
            getattr(vat_lookup, 'processing_notes', '') if vat_lookup else '',
            getattr(vat_lookup, 'proxy_used', '') if vat_lookup else '',
            vat_lookup.created_at.strftime('%Y-%m-%d %H:%M:%S') if vat_lookup and vat_lookup.created_at else ''
        ]
        
        # Write row data to sheet
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num).value = value
    
    # Auto-adjust column widths
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)


def create_website_hunting_sheet(workbook: Workbook, companies) -> None:
    """
    Create Website Hunting sheet with all website hunting data.
    
    Args:
        workbook: The Excel workbook to add sheet to
        companies: QuerySet of CompanyNumber objects with related data
    """
    # Create worksheet
    sheet = workbook.create_sheet("Website Hunting")
    
    # Define headers
    headers = [
        'Company Number',
        'Company Name (from Companies House)',
        'Domains Found (SERP)',
        'Domains Found Count',
        'Ranked Domains (with Scores)',
        'Ranked Domains Count',
        'Best Scoring Domain',
        'Best Score',
        'SERP Status',
        'Crawl Status',
        'Processing Notes',
        'Approved Domain',
        'Approved by Human',
        'Created At'
    ]
    
    # Add headers to sheet
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6FF", end_color="E6E6FF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row_num, company in enumerate(companies, 2):
        try:
            hunting_data = company.website_hunting_result
        except ObjectDoesNotExist:
            hunting_data = None
        try:
            house_data = company.house_data
        except ObjectDoesNotExist:
            house_data = None
        
        # Parse domains found and ranked domains
        domains_found_str = ''
        domains_found_count = 0
        ranked_domains_str = ''
        ranked_domains_count = 0
        best_domain = ''
        best_score = ''
        
        if hunting_data:
            # Handle domains found
            if hunting_data.domains_found:
                domains_found_count = len(hunting_data.domains_found)
                domains_found_str = '; '.join(hunting_data.domains_found) if isinstance(hunting_data.domains_found, list) else str(hunting_data.domains_found)
            
            # Handle ranked domains
            if hunting_data.ranked_domains:
                ranked_domains_count = len(hunting_data.ranked_domains)
                
                # Create readable string with domain and score
                if isinstance(hunting_data.ranked_domains, list):
                    domain_scores = []
                    best_score_val = 0
                    
                    for domain_data in hunting_data.ranked_domains:
                        if isinstance(domain_data, dict):
                            domain = domain_data.get('domain', 'Unknown')
                            score = domain_data.get('score', 0)
                            domain_scores.append(f"{domain} ({score})")
                            
                            # Track best scoring domain
                            if score > best_score_val:
                                best_score_val = score
                                best_domain = domain
                                best_score = str(score)
                    
                    ranked_domains_str = '; '.join(domain_scores)
                else:
                    ranked_domains_str = str(hunting_data.ranked_domains)
        
        # Prepare row data with fallback to empty strings
        row_data = [
            company.company_number,
            getattr(house_data, 'company_name', '') if house_data else '',
            domains_found_str,
            domains_found_count,
            ranked_domains_str,
            ranked_domains_count,
            best_domain,
            best_score,
            getattr(hunting_data, 'serp_status', '') if hunting_data else '',
            getattr(hunting_data, 'crawl_status', '') if hunting_data else '',
            getattr(hunting_data, 'processing_notes', '') if hunting_data else '',
            getattr(hunting_data, 'approved_domain', '') if hunting_data else '',
            'Yes' if hunting_data and hunting_data.approved_by_human else 'No',
            hunting_data.created_at.strftime('%Y-%m-%d %H:%M:%S') if hunting_data and hunting_data.created_at else ''
        ]
        
        # Write row data to sheet
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num).value = value
    
    # Auto-adjust column widths
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)


def create_website_contact_extraction_sheet(workbook: Workbook, companies) -> None:
    """
    Create Website Contact Extraction sheet with all contact extraction data.
    
    Args:
        workbook: The Excel workbook to add sheet to
        companies: QuerySet of CompanyNumber objects with related data
    """
    # Create worksheet
    sheet = workbook.create_sheet("Website Contact Extraction")
    
    # Define headers
    headers = [
        'Company Number',
        'Company Name (from Companies House)',
        'Domain Searched',
        'Phone Numbers',
        'Phone Numbers Count',
        'Email Addresses',
        'Email Addresses Count',
        'Social Media Links',
        'Social Media Count',
        'Facebook Links',
        'Instagram Links',
        'LinkedIn Links',
        'Total Contacts Found',
        'Pages Crawled',
        'Status',
        'Processing Notes',
        'Created At'
    ]
    
    # Add headers to sheet
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E0F8F8", end_color="E0F8F8", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row_num, company in enumerate(companies, 2):
        try:
            contact_data = company.website_contact_lookup
        except ObjectDoesNotExist:
            contact_data = None
        try:
            house_data = company.house_data
        except ObjectDoesNotExist:
            house_data = None
        
        # Initialize variables
        phone_numbers_str = ''
        phone_count = 0
        email_addresses_str = ''
        email_count = 0
        social_media_str = ''
        social_count = 0
        facebook_links = ''
        instagram_links = ''
        linkedin_links = ''
        total_contacts = 0
        
        if contact_data:
            # Handle phone numbers
            if contact_data.phone_numbers and isinstance(contact_data.phone_numbers, list):
                phone_count = len(contact_data.phone_numbers)
                phone_numbers_str = '; '.join(contact_data.phone_numbers)
            
            # Handle email addresses
            if contact_data.email_addresses and isinstance(contact_data.email_addresses, list):
                email_count = len(contact_data.email_addresses)
                email_addresses_str = '; '.join(contact_data.email_addresses)
            
            # Handle social media links
            if contact_data.social_media_links and isinstance(contact_data.social_media_links, dict):
                social_media_parts = []
                
                # Facebook links
                facebook = contact_data.social_media_links.get('facebook', [])
                if isinstance(facebook, list) and facebook:
                    facebook_links = '; '.join(facebook)
                    social_media_parts.append(f"Facebook: {facebook_links}")
                    social_count += len(facebook)
                elif isinstance(facebook, str) and facebook.strip():
                    facebook_links = facebook
                    social_media_parts.append(f"Facebook: {facebook}")
                    social_count += 1
                
                # Instagram links
                instagram = contact_data.social_media_links.get('instagram', [])
                if isinstance(instagram, list) and instagram:
                    instagram_links = '; '.join(instagram)
                    social_media_parts.append(f"Instagram: {instagram_links}")
                    social_count += len(instagram)
                elif isinstance(instagram, str) and instagram.strip():
                    instagram_links = instagram
                    social_media_parts.append(f"Instagram: {instagram}")
                    social_count += 1
                
                # LinkedIn links
                linkedin = contact_data.social_media_links.get('linkedin', [])
                if isinstance(linkedin, list) and linkedin:
                    linkedin_links = '; '.join(linkedin)
                    social_media_parts.append(f"LinkedIn: {linkedin_links}")
                    social_count += len(linkedin)
                elif isinstance(linkedin, str) and linkedin.strip():
                    linkedin_links = linkedin
                    social_media_parts.append(f"LinkedIn: {linkedin}")
                    social_count += 1
                
                social_media_str = ' | '.join(social_media_parts)
            
            # Calculate total contacts
            total_contacts = phone_count + email_count + social_count
        
        # Prepare row data with fallback to empty strings
        row_data = [
            company.company_number,
            getattr(house_data, 'company_name', '') if house_data else '',
            getattr(contact_data, 'domain_searched', '') if contact_data else '',
            phone_numbers_str,
            phone_count,
            email_addresses_str,
            email_count,
            social_media_str,
            social_count,
            facebook_links,
            instagram_links,
            linkedin_links,
            total_contacts,
            getattr(contact_data, 'pages_crawled', 0) if contact_data else 0,
            getattr(contact_data, 'status', 'NOT_PROCESSED') if contact_data else 'NOT_PROCESSED',
            getattr(contact_data, 'processing_notes', '') if contact_data else '',
            contact_data.created_at.strftime('%Y-%m-%d %H:%M:%S') if contact_data and contact_data.created_at else ''
        ]
        
        # Write row data to sheet
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num).value = value
    
    # Auto-adjust column widths
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)


def create_linkedin_profile_discovery_sheet(workbook: Workbook, companies) -> None:
    """
    Create LinkedIn Profile Discovery sheet with all LinkedIn lookup data.
    
    Args:
        workbook: The Excel workbook to add sheet to
        companies: QuerySet of CompanyNumber objects with related data
    """
    # Create worksheet
    sheet = workbook.create_sheet("LinkedIn Profile Discovery")
    
    # Define headers
    headers = [
        'Company Number',
        'Company Name (from Companies House)',
        'Company URLs',
        'Company URLs Count',
        'Employee URLs',
        'Employee URLs Count',
        'Best Company Profile URL',
        'Best Company Profile Score',
        'Total Results Found',
        'Total LinkedIn Profiles',
        'Search Query',
        'Search Status',
        'Processing Notes',
        'Approved Domain Used',
        'Created At'
    ]
    
    # Add headers to sheet
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Add data rows
    for row_num, company in enumerate(companies, 2):
        try:
            linkedin_data = company.linkedin_lookup
        except ObjectDoesNotExist:
            linkedin_data = None
        try:
            house_data = company.house_data
        except ObjectDoesNotExist:
            house_data = None
        
        # Initialize variables
        company_urls_str = ''
        company_urls_count = 0
        employee_urls_str = ''
        employee_urls_count = 0
        best_company_url = ''
        best_company_score = ''
        total_profiles = 0
        
        if linkedin_data:
            # Handle company URLs
            if linkedin_data.company_urls and isinstance(linkedin_data.company_urls, list):
                company_urls_count = len(linkedin_data.company_urls)
                
                # Create readable string with URL and score
                if company_urls_count > 0:
                    company_url_parts = []
                    best_score = 0
                    
                    for company_profile in linkedin_data.company_urls:
                        if isinstance(company_profile, dict):
                            url = company_profile.get('url', 'Unknown URL')
                            score = company_profile.get('score', 0)
                            company_url_parts.append(f"{url} ({score})")
                            
                            # Track best scoring company profile
                            if score > best_score:
                                best_score = score
                                best_company_url = url
                                best_company_score = str(score)
                        elif isinstance(company_profile, str):
                            company_url_parts.append(company_profile)
                    
                    company_urls_str = '; '.join(company_url_parts)
            
            # Handle employee URLs
            if linkedin_data.employee_urls and isinstance(linkedin_data.employee_urls, list):
                employee_urls_count = len(linkedin_data.employee_urls)
                
                # Create readable string
                if employee_urls_count > 0:
                    employee_url_parts = []
                    
                    for employee_profile in linkedin_data.employee_urls:
                        if isinstance(employee_profile, dict):
                            url = employee_profile.get('url', 'Unknown URL')
                            score = employee_profile.get('score', 0)
                            employee_url_parts.append(f"{url} ({score})")
                        elif isinstance(employee_profile, str):
                            employee_url_parts.append(employee_profile)
                    
                    employee_urls_str = '; '.join(employee_url_parts)
            
            # Calculate total profiles
            total_profiles = company_urls_count + employee_urls_count
        
        # Prepare row data with fallback to empty strings
        row_data = [
            company.company_number,
            getattr(house_data, 'company_name', '') if house_data else '',
            company_urls_str,
            company_urls_count,
            employee_urls_str,
            employee_urls_count,
            best_company_url,
            best_company_score,
            getattr(linkedin_data, 'total_results_found', 0) if linkedin_data else 0,
            total_profiles,
            getattr(linkedin_data, 'search_query', '') if linkedin_data else '',
            getattr(linkedin_data, 'search_status', 'NOT_PROCESSED') if linkedin_data else 'NOT_PROCESSED',
            getattr(linkedin_data, 'processing_notes', '') if linkedin_data else '',
            getattr(linkedin_data, 'approved_domain_used', '') if linkedin_data else '',
            linkedin_data.created_at.strftime('%Y-%m-%d %H:%M:%S') if linkedin_data and linkedin_data.created_at else ''
        ]
        
        # Write row data to sheet
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num).value = value
    
    # Auto-adjust column widths
    for column_cells in sheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(length + 2, 50)


def get_full_export_summary(campaign: Campaign) -> dict:
    """
    Get a summary of the full export data for a campaign.
    
    Args:
        campaign: The campaign to get summary for
        
    Returns:
        Dictionary with summary statistics
    """
    companies = CompanyNumber.objects.filter(campaign=campaign).select_related('house_data', 'vat_lookup', 'website_hunting_result', 'website_contact_lookup', 'linkedin_lookup')
    
    total_companies = companies.count()
    with_house_data = companies.filter(house_data__isnull=False).count()
    with_vat_data = companies.filter(vat_lookup__isnull=False).count()
    with_hunting_data = companies.filter(website_hunting_result__isnull=False).count()
    with_contact_data = companies.filter(website_contact_lookup__isnull=False).count()
    with_linkedin_data = companies.filter(linkedin_lookup__isnull=False).count()
    
    # Count website hunting specific metrics
    domains_found_count = 0
    approved_domains_count = 0
    for company in companies:
        try:
            hunting_data = company.website_hunting_result
            if hunting_data.domains_found and len(hunting_data.domains_found) > 0:
                domains_found_count += 1
            if hunting_data.approved_by_human and hunting_data.approved_domain:
                approved_domains_count += 1
        except ObjectDoesNotExist:
            continue
    
    # Count contact extraction specific metrics
    companies_with_emails = 0
    companies_with_phones = 0
    companies_with_social = 0
    companies_with_any_contact = 0
    total_emails_found = 0
    total_phones_found = 0
    total_social_found = 0
    
    for company in companies:
        try:
            contact_data = company.website_contact_lookup
            has_any_contact = False
            
            # Check emails
            if contact_data.email_addresses and len(contact_data.email_addresses) > 0:
                companies_with_emails += 1
                total_emails_found += len(contact_data.email_addresses)
                has_any_contact = True
            
            # Check phones
            if contact_data.phone_numbers and len(contact_data.phone_numbers) > 0:
                companies_with_phones += 1
                total_phones_found += len(contact_data.phone_numbers)
                has_any_contact = True
            
            # Check social media
            if contact_data.social_media_links:
                social_count = 0
                for _, links in contact_data.social_media_links.items():
                    if isinstance(links, list):
                        social_count += len(links)
                    elif isinstance(links, str) and links.strip():
                        social_count += 1
                
                if social_count > 0:
                    companies_with_social += 1
                    total_social_found += social_count
                    has_any_contact = True
            
            if has_any_contact:
                companies_with_any_contact += 1
        except ObjectDoesNotExist:
            continue
    
    # Count LinkedIn specific metrics
    companies_with_company_profiles = 0
    companies_with_employee_profiles = 0
    companies_with_any_profiles = 0
    total_company_profiles_found = 0
    total_employee_profiles_found = 0
    
    for company in companies:
        try:
            linkedin_data = company.linkedin_lookup
            has_any_profile = False
            
            # Check company profiles
            if linkedin_data.company_urls and len(linkedin_data.company_urls) > 0:
                companies_with_company_profiles += 1
                total_company_profiles_found += len(linkedin_data.company_urls)
                has_any_profile = True
            
            # Check employee profiles
            if linkedin_data.employee_urls and len(linkedin_data.employee_urls) > 0:
                companies_with_employee_profiles += 1
                total_employee_profiles_found += len(linkedin_data.employee_urls)
                has_any_profile = True
            
            if has_any_profile:
                companies_with_any_profiles += 1
        except ObjectDoesNotExist:
            continue
    
    return {
        'total_companies': total_companies,
        'companies_house_data': with_house_data,
        'vat_lookup_data': with_vat_data,
        'website_hunting_data': with_hunting_data,
        'website_contact_data': with_contact_data,
        'linkedin_data': with_linkedin_data,
        'domains_found_count': domains_found_count,
        'approved_domains_count': approved_domains_count,
        'companies_with_emails': companies_with_emails,
        'companies_with_phones': companies_with_phones,
        'companies_with_social': companies_with_social,
        'companies_with_any_contact': companies_with_any_contact,
        'companies_with_linkedin_company_profiles': companies_with_company_profiles,
        'companies_with_linkedin_employee_profiles': companies_with_employee_profiles,
        'companies_with_any_linkedin_profiles': companies_with_any_profiles,
        'total_emails_found': total_emails_found,
        'total_phones_found': total_phones_found,
        'total_social_found': total_social_found,
        'total_company_profiles_found': total_company_profiles_found,
        'total_employee_profiles_found': total_employee_profiles_found,
        'total_linkedin_profiles_found': total_company_profiles_found + total_employee_profiles_found,
        'completion_percentage': round((with_house_data / total_companies) * 100, 1) if total_companies > 0 else 0,
        'sheets_included': ['Companies House Lookup', 'VAT Lookup', 'Website Hunting', 'Website Contact Extraction', 'LinkedIn Profile Discovery'],
        'export_format': 'Excel (.xlsx)'
    }